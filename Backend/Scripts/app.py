import shutil
from flask import Flask, request, jsonify,send_file, send_from_directory,make_response
from flask_cors import CORS
from PyPDF2 import PdfMerger,PdfWriter,PdfReader
import io
import PIL.Image
import cv2
import os
import atexit
from docx import Document
import win32com.client as win32
from docx2pdf import convert
from pdf2docx import parse
from typing import Tuple
import tempfile
from docx2pdf import convert
import pythoncom
import img2pdf
from pptx import Presentation
import comtypes.client
from werkzeug.utils import secure_filename
from fpdf import FPDF
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import zipfile
import tempfile
from flask import Flask, request, send_file
from pdftabextract.common import read_xml, parse_pages
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import pandas as pd
import pdfkit
import tabula
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate,Table, TableStyle
from reportlab.lib import colors
import camelot
from werkzeug.utils import secure_filename
from pdf2image import convert_from_path
from pdf2pptx import convert_pdf2pptx
import tempfile
import json
import requests
import aspose.slides as slides
import os
import speech_recognition as sr
from pydub import AudioSegment
import PyPDF2
import pdfkit
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from werkzeug.utils import secure_filename
import youtube_dl
from fpdf import FPDF
from reportlab.lib.units import inch
from moviepy.editor import *
from moviepy.editor import *
from pytube import YouTube
import pytube
import subprocess


from moviepy.editor import VideoFileClip




app = Flask(__name__)
CORS(app)

ffmpeg_path =  r"C:\Program Files (x86)\FormatFactory\ffmpeg.exe"
config = pdfkit.configuration(wkhtmltopdf= r"C:\wkhtmltopdf\bin\wkhtmltopdf.exe")


# This Are The Components of pdf like merge , split , add page number , protect and unlock the pdf

# MERGER THE PDFS

@app.route('/merge', methods=['POST'])
def merge_pdf():
    merger = PdfMerger()

    for file in request.files.getlist('files'):
        pdf = PdfReader(file)
        merger.append(pdf)

    output = io.BytesIO()
    merger.write(output)
    output.seek(0)

    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name='merged.pdf')



#This Module is Adding page Number In the pdf
@app.route('/PageNumberAdd', methods=['POST'])
def add_page_numbers():
    # Get the uploaded PDF file from the request
    file = request.files['file']

    # Get user's selected position, size, and color
    position = request.form['position']
    size = int(request.form['size'])  # Convert size to an integer
    color = request.form['color']

    # Read the PDF file
    pdf_reader = PdfReader(file)
    total_pages = len(pdf_reader.pages)

    # Create a new PDF writer
    pdf_writer = PdfWriter()

    # Add page numbers to each page
    for i, page in enumerate(pdf_reader.pages):
        page_number = i + 1

        # Create a canvas to write the page number
        packet = BytesIO()
        can = canvas.Canvas(packet, pagesize=letter)
        can.setFont("Helvetica", size)  # Set the font size based on the user's choice
        # Adjust the font name and size as needed

        # Set the color based on the user's choice
        if color == 'red':
            can.setFillColorRGB(1, 0, 0)  # Red color
        elif color == 'white':
            can.setFillColorRGB(1, 1, 1)  # White color
        elif color == 'black':
            can.setFillColorRGB(0, 0, 0)  # Black color
        elif color == 'yellow':
            can.setFillColorRGB(1, 1, 0)  # Yellow color

        # Set the position based on the user's choice
        if position == 'top-left':
            x = 10
            y = float(page.mediabox[3]) - 20  # Convert to float
        elif position == 'top-center':
            x = page.mediabox[2] / 2
            y = float(page.mediabox[3]) - 20  # Convert to float
        elif position == 'top-right':
            x = float(page.mediabox[2]) - 50  # Convert to float
            y = float(page.mediabox[3]) - 20  # Convert to float
        elif position == 'bottom-left':
            x = 10
            y = 10
        elif position == 'bottom-center':
            x = page.mediabox[2] / 2
            y = 10
        elif position == 'bottom-right':
            x = float(page.mediabox[2]) - 50  # Convert to float
            y = 10

        # Write the page number on the canvas
        can.drawString(x, y, str(page_number))

        # Move the canvas to the beginning of the packet
        can.save()
        packet.seek(0)

        # Create a new PDF page with the page number
        new_page = PdfReader(packet).pages[0]

        # Merge the modified page with the original page
        page.merge_page(new_page, expand=False)

        # Add the modified page to the PDF writer
        pdf_writer.add_page(page)

    # Create a new file name for the modified PDF
    filename = 'modified_' + file.filename
    output_filename = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs\{}.pdf".format(filename))

    # Save the modified PDF file
    with open(output_filename, 'wb') as output_file:
        pdf_writer.write(output_file)

    file_path = r"C:\Users\HP\Desktop\ConverterArea\Backend/PDF_Outputs\{}.pdf".format(filename)

    # Return the file URL for download
    return send_file(file_path, as_attachment=True)


# This Module is the Protect the pdf by adding Password

@app.route('/Protect_Pdf', methods=['POST'])
def protect_pdf():
    password = request.form["password"]
    file = request.files["file"]

    # Read the PDF file
    pdf_reader = PdfReader(file)
    
    # Create a new PDF writer
    pdf_writer = PdfWriter()
    
    # Encrypt each page with the provided password
    for page in pdf_reader.pages:
        pdf_writer.add_page(page)
        pdf_writer.encrypt(password)

    # Create a temporary file to save the protected PDF
    with tempfile.NamedTemporaryFile(delete=False) as temp_pdf:
        pdf_writer.write(temp_pdf)

    # Return the protected PDF file as a response
    return send_file(temp_pdf.name, as_attachment=True)



# This Module is the Split The Pdf FIle
@app.route('/split_pdf', methods=['POST'])
def split_pdf():
    file = request.files['file']
    if not file or not file.filename.lower().endswith('.pdf'):
        return "Invalid PDF file.", 400

    try:
        # Define the output path
        output_path = r'C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs'
        os.makedirs(output_path, exist_ok=True)

        # Save the uploaded PDF file
        file_path = os.path.join(output_path, file.filename)
        file.save(file_path)

        # Read the PDF file
        pdf = PdfReader(file_path)
        # print("read")
        # Create a folder to store individual pages
        pages_folder = os.path.join(output_path, 'pages')
        os.makedirs(pages_folder, exist_ok=True)

        # Split the PDF into individual pages
        for page_num in range(len(pdf.pages)):
            output_file = os.path.join(pages_folder, f"page_{page_num + 1}.pdf")
            pdf_writer = PdfWriter()
            # print("write")
            pdf_writer.add_page(pdf.pages[page_num])
           
            with open(output_file, "wb") as output:
                pdf_writer.write(output)

        # Create a zip file
        zip_file_path = os.path.join(output_path, 'split_pages.zip')
        with zipfile.ZipFile(zip_file_path, "w") as zip_obj:
            # Add all the individual pages to the zip file
            for root, _, files in os.walk(pages_folder):
                for file in files:
                    zip_obj.write(os.path.join(root, file), file)
        # print("1")
        # Clean up - remove the pages folder
        for root, dirs, files in os.walk(pages_folder, topdown=False):
            for file in files:
                os.remove(os.path.join(root, file))
            for dir in dirs:
                os.rmdir(os.path.join(root, dir))
        os.rmdir(pages_folder)
        # print("2")
        # Return the path to the zip file
        file_result=r"C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs\split_pages.zip"
        return send_file(file_result, as_attachment=True)
    
    except Exception as e:
        return str(e), 500



# This Module is the Convert the text file into Handwritten file

@app.route("/add_watermark", methods=["POST"])
@app.route("/convert_youtube_to_slides", methods=["POST"])
def convert_youtube_to_slides():
    video_url = request.form.get("video_url")
    start_time = float(request.form.get("start_time"))
    end_time = float(request.form.get("end_time"))
    output_format = request.form.get("output_format")

    try:
        # Download YouTube video
        youtube = YouTube(video_url)
        video = youtube.streams.get_highest_resolution()
        video_path = "video.mp4"
        video.download(output_path="backend/videos", filename="video")

        # Extract slides from video
        video_clip = VideoFileClip(video_path).subclip(start_time, end_time)
        slides_path = "slides"
        slides = video_clip.write_images_sequence(f"{slides_path}/slide%03d.png")

        if output_format == "pdf":
            # Convert slides to PDF
            pdf_path = f"backend/pdf/slides.pdf"
            video_clip.write_videofile(pdf_path, codec="png", audio=False, ffmpeg_executable=ffmpeg_path)
            return send_file(pdf_path, mimetype="application/pdf", as_attachment=True)
        elif output_format == "zip":
            # Convert slides to images and create ZIP file
            images_path = f"backend/images"
            video_clip.write_images_sequence(f"{images_path}/slide%03d.png")
            zip_path = f"backend/slides.zip"
            shutil.make_archive(zip_path, "zip", images_path)
            return send_file(zip_path, mimetype="application/zip", as_attachment=True)
    except Exception as e:
        print("An error occurred:", str(e))
        return "Error converting YouTube video to slides", 500







#**********************************************************************************************************************************************
#**********************************************************************************************************************************************
#**********************************************************************************************************************************************
#**********************************************************************************************************************************************
#***CONVERT FROM PDF FILES LIKE PDFTOEXCLE.PDFTOJPG,PDFTOPOWERPOINT,PDFTOWORD***#
@app.route('/pdf_to_excel', methods=['POST'])




#CONVERT PDF TO WORD FILE
@app.route('/pdf_to_word', methods=['POST'])
def handle_pdf_to_word():
    if 'file' not in request.files:
        return "No file uploaded", 400

    file = request.files['file']
    input_file = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\Temp", file.filename)
    name = secure_filename(file.filename)
    output_file = r"C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs\{}.docx".format(name)
    file.save(input_file)

    convert_pdf2docs(input_file, output_file)

    return send_file(output_file, as_attachment=True)

def convert_pdf2docs(input_file: str, output_file: str, pages: Tuple = None):
    if pages:
        pages = [int(i) for i in list(pages) if i.isnumeric()]
    result = parse(pdf_file=input_file, docx_file=output_file, pages=pages)
    return result
    
@app.route('/convert_pdf_to_images', methods=['POST'])
def convert_pdf_to_images():
    # Save the uploaded PDF file
    pdf_file = request.files['file']
    pdf_path = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\Temp", 'input.pdf')
    pdf_file.save(pdf_path)

    # Convert PDF to images
    images = convert_from_path(pdf_path)

    # Save the images and create a list of their file paths
    image_paths = []
    for i, image in enumerate(images):
        image_path = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs", f'image_{i+1}.jpg')
        image.save(image_path, 'JPEG')
        image_paths.append(image_path)

    zip_path = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs", 'converted_images.zip')
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        for image_path in image_paths:
            zip_file.write(image_path, os.path.basename(image_path))

    # Send the ZIP archive back to the frontend for download
    return send_file(zip_path, as_attachment=True)


#***********************************************************************************************************************************************#
#**********************************************************************************************************************************************
#**********************************************************************************************************************************************
#**********************************************************************************************************************************************
#**********************************************************************************************************************************************
#***CONVERT TO PDF FILES ***#
# This Module is the Convert the Excel into Pdf file
@app.route('/excel_to_pdf', methods=['POST'])
def convert_endpoint():
    # Check if the request contains a file
    if 'file' not in request.files:
        return "No file found", 400

    file = request.files['file']
    pdf_filename = convert_excel_to_pdf(file)
    return send_file(pdf_filename, as_attachment=True)

def convert_excel_to_pdf(file):
    # Load the Excel file using pandas
    df = pd.read_excel(file)

    # Get the formatting information from the Excel file
    wb = load_workbook(file)
    ws = wb.active
    max_col = ws.max_column
    cell_formats = [ws.cell(row=1, column=col).number_format for col in range(1, max_col + 1)]

    # Create a list to hold the table data
    data = [df.columns.tolist()] + df.values.tolist()

    # Generate the PDF file
    pdf_filename = 'output.pdf'
    doc = SimpleDocTemplate(pdf_filename, pagesize=letter)

    # Define the table style with formatting information
    table_style = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEADING', (0, 0), (-1, -1), 12),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black)
    ])

    elements = []
    table = Table(data, style=table_style)

    # Apply formatting information to the table cells
    for row in range(len(data)):
        for col in range(len(data[0])):
            cell_format = cell_formats[col]
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (col, row), (col, row), colors.black),
                ('FONTNAME', (col, row), (col, row), 'Helvetica'),
                ('FONTSIZE', (col, row), (col, row), 10),
                ('LEADING', (col, row), (col, row), 12),
                ('BACKGROUND', (col, row), (col, row), colors.white),  # Set the background to white
                ('ALIGN', (col, row), (col, row), 'LEFT'),
                ('VALIGN', (col, row), (col, row), 'MIDDLE'),
            ]))

    elements.append(table)

    doc.build(elements)

    return pdf_filename



# This Module is the convert the Html file into Pdf
@app.route('/Html_to_pdf', methods=['POST'])
def convert_html_to_pdf():
    # Get the URL from the request JSON data
    url = request.json['url']

    # Set the output PDF file path
    output_path = 'output.pdf'

    # Convert HTML to PDF using pdfkit library
    pdfkit.from_url(url, output_path)



@app.route('/Image_to_pdf', methods=['POST'])
def convert_images_to_pdf():
    # Get the uploaded images
    images = request.files.getlist('files')

    # Create a list to store image paths
    image_paths = []

    # Save the uploaded images and collect their paths
    for i, image in enumerate(images):
        filename = f"input_{i+1}.jpg"  # Adjust the filename format if desired
        file_path = os.path.join(r'C:\Users\HP\Desktop\ConverterArea\Backend\Temp', filename)
        image.save(file_path)
        image_paths.append(file_path)

    # Convert images to PDF
    output_path = os.path.join(r'C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs', "output.pdf")
    with open(output_path, "wb") as f:
        f.write(img2pdf.convert(image_paths, rotation=img2pdf.Rotation.ifvalid))

    # Send the converted PDF file back to the frontend
    return send_file(output_path, as_attachment=True)




@app.route('/PPt_to_pdf', methods=['POST'])
def convert_to_pdf():
    # Get the uploaded PowerPoint file
    ppt_file = request.files['file']

    # Save the uploaded PowerPoint file
    input_path = os.path.join(r'C:\Users\HP\Desktop\ConverterArea\Backend\Temp', 'input.pptx')
    ppt_file.save(input_path)

    # Convert PowerPoint to PDF
    output_path = os.path.join(r'C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs', 'output.pdf')
    pythoncom.CoInitialize()

    # Create PowerPoint application object
    powerpoint = comtypes.client.CreateObject("Powerpoint.Application")

    # Open the PPTX file
    presentation = powerpoint.Presentations.Open(input_path)

    # Save the presentation as PDF
    presentation.ExportAsFixedFormat(output_path, 2)  # 2 represents PDF format

    # Close the presentation and quit PowerPoint application
    presentation.Close()
    powerpoint.Quit()

    # Send the converted PDF file back to the frontend
    return send_file(output_path, as_attachment=True)




# This Module is the Convert the text file into Handwritten file

@app.route('/text_to_handwritten', methods=['POST'])
def text_to_handwritten():
   try:
      file = request.files["file"]
      scale_percent =int( request.form.get("value"))
      font = request.form.get("font_style")
      filePath = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\Temp", file.filename)
      file.save(filePath)
      file_extension = os.path.splitext(file.filename)[1]

      if file_extension == ".docx":
      # Convert DOCX to TXT
        txt_filename = os.path.splitext(file.filename)[0] + ".txt"
        txt_filepath = os.path.join(r"C:\Users\HP\Desktop\ConverterArea\Backend\Temp", txt_filename)

        doc = Document(filePath)
        full_text = []
        for paragraph in doc.paragraphs:
            full_text.append(paragraph.text)

     # Save the TXT file
        with open(txt_filepath, "w", encoding="utf-8") as txt_file:
            txt_file.write("\n".join(full_text))
  
      # Update the file path to the converted TXT file
        filePath = txt_filepath
        
      os.makedirs("../Output", exist_ok=True)
      os.makedirs("../PDF_Outputs", exist_ok=True)
      global font_style,background,SheetWidth,margin,lineMargin,lineGap,pageNum,allowedCharacters,wordsPerLine,FontType,x,y
      font_style = font
      background = PIL.Image.open("../Fonts/myfont/a4.jpg")
      SheetWidth = background.width
      margin = 115
      lineMargin = 115
      allowedCharacters = '''ABCDEFGHIJKLMNOPQRSTUVWXYZ 
                        abcdefghijklmnopqrstuvwxyz 
                        #:,.?-!{}()[]'<>+=%^$@_;1234567890 "'''

      wordsPerLine = 80
      maxLenPerPage = 3349
      pageNum = 1 

      if font_style.lower() == "uv":
         lineGap = 120
      elif font_style.lower() == "rajat" or font_style.lower() == "swagat":
         lineGap = 165
      elif font_style.lower() == "piyush":
         lineGap = 144
      else:
        lineGap = 150

      FontType = "../Fonts/{}_font/".format(font_style)
      x, y = margin + 20, margin + lineGap

      print("Starting.")
         # Get the file, value, and font style from the request
      file = open(filePath , "r")
      content = file.read()

      l = len(content)
      content = content.split("\n")

      print("Text Reading Completed.")
      for i in range(len(content)):
            writeByLine(content[i])
            newLine()
      print("Saved Page: ", pageNum)

      background.save("../Output/{}_output_{}.png".format(font_style, pageNum))

      ImagesPath = [
            "../Output/{}_output_{}.png".format(font_style, page)
            for page in range(1, pageNum + 1)
        ]

      print("Adding lines.")

      for path in ImagesPath:
          img = cv2.imread(path, cv2.IMREAD_COLOR)
          x  = 0
          y = 228

          cv2.line(img, (lineMargin - 20, 0), (lineMargin - 20, 3508), (0, 0, 0), 3)
          cv2.line(img, (x, y), (x + 2480, y), (0, 0, 0), 2)

          while y <= 3349:
             cv2.line(img, (x, y), (x + 2480, y), (0, 0, 0), 2)
             y += lineMargin

          width = int(img.shape[1] * scale_percent / 100)
          height = int(img.shape[0] * scale_percent / 100)
          dim = (width, height)

          mimage = cv2.resize(img, dim, interpolation=cv2.INTER_NEAREST)
          cv2.imwrite(path, mimage)

      height, width = PIL.Image.open(ImagesPath[0]).size
      pdf = FPDF(unit="pt", format=(height, width))

      for i in range(0, pageNum):
         pdf.add_page()
         pdf.image(ImagesPath[i], 0, 0)

      print("Saving the pdf.")
      pdf_name = "../PDF_Outputs/{}.pdf".format(font_style)
      pdf.output(pdf_name, "F")

      print("Removing unnecessary files.")
      for path in ImagesPath:
            os.remove(path)

      print("Done.")
      print("Find your output at " + pdf_name + ".")
      file_path = r"C:\Users\HP\Desktop\ConverterArea\Backend/PDF_Outputs\{}.pdf".format(font_style)
      return send_file(file_path, as_attachment=True)
   
   except Exception as e:
        pass
        print(e, "Try again!")
        return jsonify({"error": str(e)}), 500
       
   
def space():
      global x, y

      space = PIL.Image.open("../Fonts/myfont/space.png")
      width = space.width
      x += width
      background.paste(space, (x, y))
      del space


def newLine():
       global x, y

       x = margin + 20
       y += margin


def writeAlphabet(path):
        global x, y

        letter = PIL.Image.open(path)
        background.paste(letter, (x, y))
        x += letter.width


def check_pageExceed():
        global font_style, pageNum, background, x, y, margin, lineGap

        if y >= 3100:
             background.save("../Output/{}_output_{}.png".format(font_style, pageNum))
             print("Saved Page: ", pageNum)
             bg = PIL.Image.open("../Fonts/myfont/a4.jpg")
             background = bg
             x, y = margin, margin + lineGap
             pageNum += 1


wasDQ = False

def ProcessNwrite(word):
        global x, y, background, pageNum, font_style, margin, lineGap, wasDQ

        if x > SheetWidth - wordsPerLine * len(word):
             newLine()

        check_pageExceed()

        path = FontType
        for letter in word:
            if letter in allowedCharacters:
                if letter.isupper():
                   path += "upper/{}".format(letter)
                elif letter.islower():
                   path += "lower/{}".format(letter)
                elif letter.isnumeric():
                   path += "digits/{}".format(letter)
                else:
                   path += "symbols/"
                   if letter == ",":
                       path += "comma"
                   elif letter == ".":
                        path += "fullstop"
                   elif letter == "!":
                        path += "exclamation"
                   elif letter == "-":
                        path += "hiphen"
                   elif letter == "#":
                        path += "hash"
                   elif letter == "?":
                        path += "question"
                   elif letter == "(":
                       path += "bracketop"
                   elif letter == ")":
                      path += "bracketcl"
                   elif letter == ":":
                      path += "colon"
                   elif letter == ";":
                      path += "semicolon"
                   elif letter == "{":
                      path += "Cbracketop"
                   elif letter == "}":
                      path += "Cbracketcl"
                   elif letter == "[":
                      path += "osb"
                   elif letter == "]":
                      path += "csb"
                   elif letter == "<":
                      path += "oab"
                   elif letter == ">":
                      path += "cab"
                   elif letter == "=":
                      path += "equals"
                   elif letter == "'":
                      path += "osq"
                   elif letter == "%":
                      path += "percent"
                   elif letter == "&":
                      path += "empersand"
                   elif letter == "$":
                      path += "dollar"
                   elif letter == "@":
                      path += "at"
                   elif letter == "*":
                      path += "asterisk"
                   elif letter == "_":
                      path += "underscore"
                   elif letter == "^":
                      path += "cap"
                   elif letter == '"' and wasDQ:
                      path += "cdq"
                      wasDQ = False
                   elif letter == '"':
                      path += "odq"
                      wasDQ = True
                path += ".png"

                writeAlphabet(path)
                path = FontType
            else:
               writeAlphabet("../Fonts/myfont/space.png")


def writeByLine(data):
        global x, y, background, pageNum, font_style
        print("Come to writeByLine")
        if data == "":
           newLine
           print("Complete if")
        else:
           print("come to else")
           data = data.split(" ")
           check_pageExceed()
           print("complete checking")

           for word in data:
               ProcessNwrite(word)
               space()

# This module is convert the word file to pdf
@app.route('/Word_to_pdf', methods=['POST'])
def convert_word_to_pdf():
    # Save the uploaded Word file
    word_file = request.files['file']
    input_path = os.path.join(r'C:\Users\HP\Desktop\ConverterArea\Backend\Temp', 'input.docx')
    word_file.save(input_path)

    # Initialize the COM library
    pythoncom.CoInitialize()

    # Convert Word file to PDF
    output_path = os.path.join(r'C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs', 'output.pdf')
    convert(input_path, output_path)

    # Send the converted PDF file back to the frontend
    return send_file('output.pdf', as_attachment=True)


@app.route('/audio_to_pdf', methods=['POST'])




def convert_audio_to_pdf(audio_path, output_path):
    try:
        # Load the audio file
        audio = AudioSegment.from_file(audio_path)

        # Convert the audio to text using SpeechRecognition
        recognizer = sr.Recognizer()
        with sr.AudioFile(audio_path) as audio_file:
            audio_data = recognizer.record(audio_file)
            text = recognizer.recognize_google(audio_data)

        # Generate the PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.multi_cell(0, 10, text)
        pdf.output(output_path)

        return True

    except Exception as e:
        print(f"Error converting audio to PDF: {str(e)}")
        return False







# Deleting Temory file
folder_path1 = r"C:\Users\HP\Desktop\ConverterArea\Backend\PDF_Outputs"
folder_path2 = r"C:\Users\HP\Desktop\ConverterArea\Backend\Temp"
def delete_files():
    delete_files_in_folder(folder_path1)
    delete_files_in_folder(folder_path2)
    
def delete_files_in_folder(folder_path):
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(f"Deleted file: {file_path}") 
atexit.register(delete_files)            
if __name__ == '__main__':
    app.run(debug=True)
    




